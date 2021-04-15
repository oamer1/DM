#!/pkg/qct/software/python/3.6.0/bin/python

""" Contains the module for handling spreadsheets

    import dm
    proj = Process()
"""

import argparse
import csv
import os
import sys
from abc import ABCMeta, abstractmethod

import openpyxl
import xlrd

import log

__all__ = ["Spreadsheet_if", "Spreadsheet_csv", "Spreadsheet_xls", "Spreadsheet_xlsx"]

LOGGER = log.getLogger(__name__)


class Spreadsheet_if(metaclass=ABCMeta):
    """Abstract class to define the interface for the Spreadsheet class

    This is the base class which defines which methods to define. This will be
    extended to support xls files, xlsx files, and csv files.

    Attributes:
        active_sheet (Sheet): Stores the specific sheet object that is open
        ss (Spreadsheet): Stores the specific spreadsheet object that is open

    Methods:
        open_ss: Open the specified spreadsheet filename
        set_active_sheet: Open the specified sheet name
        set_active_sheet_no: Open the specified sheet number
        get_row_by_number: In the active sheet, return the specified row
        get_row_count: Get the number of rows in the active sheet
        iter_rows: Generator to return all rows in the active sheet

    Examples:
        >>> ss = Spreadsheet_if.new_ss("test.csv")

    """

    def __init__(self) -> None:
        """Constructor for the Spreadsheet_if object"""
        log.call_arguments()

        self.active_sheet = None
        self.ss = None

    @abstractmethod
    def open_ss(self, fname: str) -> None:
        """Open the specified spreadsheet filename"""
        pass  # pragma: no cover

    @abstractmethod
    def set_active_sheet(self, name: str) -> None:
        """Open the specified sheet name"""
        pass  # pragma: no cover

    @abstractmethod
    def set_active_sheet_no(self, number: int = 0) -> None:
        """Open the specified sheet number"""
        pass  # pragma: no cover

    @abstractmethod
    def get_row_by_number(self, row: int) -> None:
        """In the active sheet, return the specified row"""
        return None  # pragma: no cover

    @abstractmethod
    def get_row_count(self):
        """Get the number of rows in the active sheet"""
        return None  # pragma: no cover

    def iter_rows(self, min_row: int = 0):
        """Generator to return all rows in the active sheet"""
        for row_idx in range(min_row, self.get_row_count()):
            yield self.get_row_by_number(row_idx)

    @classmethod
    def new_ss(cls, fname: str) -> "Spreadsheet_if":
        """Factory to create the proper spreadsheet object and open the spreadsheet"""
        if os.path.splitext(fname)[1] == ".xls":
            ss = Spreadsheet_xls()
        elif os.path.splitext(fname)[1] == ".csv":
            ss = Spreadsheet_csv()
        else:
            ss = Spreadsheet_xlsx()
        ss.open_ss(fname)
        return ss


class Spreadsheet_csv(Spreadsheet_if):
    """Class to handle the spreadsheet method for CSV files"""

    def __init__(self) -> None:
        """Constructor for the Spreadsheet_csv object"""
        self._rows = []
        super().__init__()

    def open_ss(self, fname: str) -> None:
        """Open the specified spreadsheet CSV files and load it into _rows"""
        with open(fname, "r") as csvfile:
            reader = csv.reader(csvfile, delimiter=",")
            self._rows = list(reader)

    def set_active_sheet(self, name: str) -> None:
        """Not implemented for CSV files"""
        # TODO - what to do here?
        pass

    def set_active_sheet_no(self, number: int = 0):
        """Not implemented for CSV files"""
        # TODO - what to do here?
        pass

    def get_row_by_number(self, row):
        """Return the row element for the specified row"""
        return self._rows[max(0, row - 1)]

    def get_row_count(self):
        """Return the number of rows in the CSV files"""
        return len(self._rows)


class Spreadsheet_xls(Spreadsheet_if):
    """Class to handle the spreadsheet method for XLS files"""

    def open_ss(self, fname: str):
        """Open the specified spreadsheet XLS using xlrd"""
        self.ss = xlrd.open_workbook(fname)
        self.set_active_sheet_no(0)

    def set_active_sheet_no(self, number: int = 0):
        """Open the specified sheet number"""
        self.active_sheet = self.ss.sheet_by_index(number)

    def set_active_sheet(self, name: str):
        """Open the specified sheet name"""
        self.active_sheet = self.ss.sheet_by_name(name)

    def get_row_by_number(self, row):
        """In the active sheet, return the specified row"""
        return self.active_sheet.row_values(row)

    def get_row_count(self):
        """Get the number of rows in the active sheet"""
        return self.active_sheet.nrows


class Spreadsheet_xlsx(Spreadsheet_if):
    """Class to handle the spreadsheet method for XLSX files"""

    def open_ss(self, fname: str) -> None:
        """Open the specified spreadsheet XLSX using openpyxl"""
        self.ss = openpyxl.load_workbook(fname, data_only=True)
        self.set_active_sheet_no(0)

    def set_active_sheet(self, name: str) -> None:
        """Open the specified sheet name"""
        self.active_sheet = self.ss[name]

    def set_active_sheet_no(self, number: int = 0):
        """Open the specified sheet number"""
        self.active_sheet = self.ss.get_sheet_by_name(self.ss.sheetnames[number])

    def get_row_by_number(self, row):
        """In the active sheet, return the specified row"""
        return self.active_sheet[str(max(1, row))]

    def get_row_count(self):
        """Get the number of rows in the active sheet"""
        return self.active_sheet.max_row


def main():
    """Main routine that is invoked when you run the script"""
    parser = argparse.ArgumentParser(
        description="Test script for the spreadsheet interface class.",
        add_help=True,
        argument_default=None,  # Global argument default
    )
    parser.add_argument("-i", "--input", help="Specify the input CSV file")
    parser.add_argument(
        "-d", "--debug", action="store_true", help="enable debug outputs"
    )
    parser.add_argument(
        "-I", "--interactive", action="store_true", help="enable an interactive session"
    )
    parser.add_argument("-t", "--test", action="store_true", help="run the doc tester")
    args = parser.parse_args()

    if args.debug:
        log.set_debug()

    if args.input:
        ss = Spreadsheet_if.new_ss(args.input)  # noqa

    if args.test:
        import doctest

        doctest.testmod()
        import IPython  # type: ignore

    elif args.interactive:
        import IPython

        IPython.embed()  # jump to ipython shell


if __name__ == "__main__":  # pragma: no cover
    log.cli_arguments()
    sys.exit(main())
